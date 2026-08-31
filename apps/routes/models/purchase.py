from datetime import datetime
import time

from ... import db
from ...database.db_workshops import Workshops
from ...database.db_suppliers import Suppliers
from ...database.db_products import Products
from ...database.db_purchases import Purchases
from ...database.db_purchase_details import PurchaseDetails

from openpyxl import load_workbook
from ...utilities.validators import role_validator, purchase_validator, excel_file_validator, purchase_excel_validator, subscription_validator
from apps.utilities.responseHelpers import *
from apps.utilities.utilities import current_timestamp
from apps.utilities.formatter import format_date


# PURCHASE MODEL CLASS ============================================================ Begin
class PurchaseModels():

    # CREATE PURCHASE ============================================================ Begin
    def create_purchase(user_role, workshop_id, datas):
        try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                return authorization_error()

            subscription_access = subscription_validator(user_role, workshop_id)

            if not subscription_access:
                return subscription_required()
            # Access Validation ---------------------------------------- Finish

            # Check Request Body ---------------------------------------- Start
            if datas is None:
                return invalid_params()

            required_data = [
                "supplier_id",
                "purchase_date",
                "purchase_details"
            ]

            for req in required_data:
                if req not in datas:
                    return parameter_error(
                        f"Missing {req} in request body."
                    )
            # Check Request Body ---------------------------------------- Finish

            # Initialize Data Input ---------------------------------------- Start
            supplier_id = datas["supplier_id"]
            purchase_date = datas["purchase_date"]
            purchase_details = datas["purchase_details"]
            # Initialize Data Input ---------------------------------------- Finish

            # Data Validation ---------------------------------------- Start
            checker_result = purchase_validator(
                supplier_id,
                purchase_date,
                purchase_details,
                workshop_id
            )

            if len(checker_result) != 0:
                return defined_error(
                    checker_result,
                    "Defined Error",
                    499
                )
            # Data Validation ---------------------------------------- Finish

            # Check Workshop ---------------------------------------- Start
            workshop = Workshops.query.filter_by(
                id=workshop_id,
                is_delete=0
            ).first()

            if not workshop:
                return not_found(
                    "Workshop could not be found."
                )
            # Check Workshop ---------------------------------------- Finish

            # Check Supplier ---------------------------------------- Start
            supplier = Suppliers.query.filter_by(
                id=supplier_id,
                workshop_id=workshop_id,
                is_delete=0
            ).first()

            if not supplier:
                return not_found(
                    "Supplier could not be found."
                )
            # Check Supplier ---------------------------------------- Finish

            # Initialize Transaction ---------------------------------------- Start
            timestamp = current_timestamp()
            total = 0
            # Initialize Transaction ---------------------------------------- Finish

            # Insert Purchase ---------------------------------------- Start
            purchase = Purchases(
                supplier_id=supplier_id,
                workshop_id=workshop_id,
                purchase_date=purchase_date,
                total=0,
                created_at=timestamp,
                updated_at=timestamp
            )

            try:
                db.session.add(purchase)
                db.session.flush()

            except Exception as e:
                db.session.rollback()
                return parameter_error(str(e))
            # Insert Purchase ---------------------------------------- Finish

            # Insert Purchase Detail ---------------------------------------- Start
            for item in purchase_details:

                product = Products.query.filter_by(
                    id=item["product_id"],
                    workshop_id=workshop_id,
                    is_delete=0
                ).first()

                if not product:
                    db.session.rollback()

                    return not_found(
                        "Product could not be found."
                    )

                quantity = int(item["quantity"])
                unit_cost = int(item["unit_cost"])
                subtotal = quantity * unit_cost

                detail = PurchaseDetails(
                    purchase_id=purchase.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    subtotal=subtotal
                )

                db.session.add(detail)

                # Update Stock
                product.stock += quantity

                # Update harga beli terakhir
                product.purchase_price = unit_cost

                # Calculate Total
                total += subtotal
            # Insert Purchase Detail ---------------------------------------- Finish

            # Update Purchase Total ---------------------------------------- Start
            purchase.total = total
            purchase.updated_at = timestamp
            # Update Purchase Total ---------------------------------------- Finish

            # Commit Transaction ---------------------------------------- Start
            try:
                db.session.commit()

            except Exception as e:
                db.session.rollback()
                return parameter_error(str(e))
            # Commit Transaction ---------------------------------------- Finish

            # Return Response ========================================
            return success(
                status_code=201
            )

        except Exception as e:
            db.session.rollback()
            return bad_request(str(e))
    # CREATE PURCHASE ============================================================ End

    # READ PURCHASE ============================================================ Begin
    def read_purchase(user_role, workshop_id):
        try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                return authorization_error()
            # Access Validation ---------------------------------------- Finish

            # Get Data ---------------------------------------- Start
            purchases = Purchases.query.filter_by(
                workshop_id=workshop_id,
                is_delete=0
            ).order_by(
                Purchases.purchase_date.desc()
            ).all()
            # Get Data ---------------------------------------- Finish

            # Initialize Data ---------------------------------------- Start
            data = []

            for purchase in purchases:

                created_at = format_date(purchase.created_at)
                updated_at = format_date(purchase.updated_at)
                deleted_at = None

                if purchase.deleted_at:
                    deleted_at = format_date(purchase.deleted_at)
                    
                purchase_date = format_date(purchase.purchase_date)

                total_item = PurchaseDetails.query.filter_by(
                    purchase_id=purchase.id
                ).count()

                data.append({
                    "id": purchase.id,
                    "supplier_id": purchase.supplier_id,
                    "supplier_name": purchase.suppliers.name if purchase.suppliers else "-",
                    "purchase_date": purchase_date,
                    "total": purchase.total,
                    "total_item": total_item,
                    "created_at": created_at,
                    "updated_at": updated_at,
                    "deleted_at": deleted_at
                })
            # Initialize Data ---------------------------------------- Finish

            # Return Response ========================================
            return success_data(
                data=data,
                status_code=200
            )

        except Exception as e:
            return bad_request(str(e))
    # READ PURCHASE ============================================================ End

    # READ PURCHASE DETAIL ============================================================ Begin
    def read_purchase_detail(user_role, workshop_id, id):
        try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                return authorization_error()
            # Access Validation ---------------------------------------- Finish

            # Check Purchase ---------------------------------------- Start
            purchase = Purchases.query.filter_by(
                id=id,
                workshop_id=workshop_id,
                is_delete=0
            ).first()

            if not purchase:
                return not_found(
                    "Purchase could not be found."
                )
            # Check Purchase ---------------------------------------- Finish

            # Initialize Detail ---------------------------------------- Start
            details = []

            for item in purchase.purchase_details:

                details.append({
                    "id": item.id,
                    "product_id": item.product_id,
                    "product_name": item.products.product_name if item.products else "-",
                    "quantity": item.quantity,
                    "unit_cost": item.unit_cost,
                    "subtotal": item.subtotal
                })
            # Initialize Detail ---------------------------------------- Finish

            # Initialize Header ---------------------------------------- Start
            purchase_date = format_date(purchase.purchase_date)

            data = {
                "id": purchase.id,
                "supplier_id": purchase.supplier_id,
                "supplier_name": purchase.suppliers.name,
                "purchase_date": purchase_date,
                "total": purchase.total,
                "details": details
            }
            # Initialize Header ---------------------------------------- Finish

            # Return Response ========================================
            return success_data(
                data=data,
                status_code=200
            )

        except Exception as e:
            return bad_request(str(e))
    # READ PURCHASE DETAIL ============================================================ End

    # UPDATE PURCHASE ============================================================ Begin
    def update_purchase(user_role, workshop_id, id, datas):
        try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                return authorization_error()
            
            subscription_access = subscription_validator(user_role, workshop_id)

            if not subscription_access:
                return subscription_required()
            # Access Validation ---------------------------------------- Finish

            # Check Request Body ---------------------------------------- Start
            if datas is None:
                return invalid_params()

            required_data = [
                "supplier_id",
                "purchase_date",
                "purchase_details"
            ]

            for req in required_data:
                if req not in datas:
                    return parameter_error(
                        f"Missing {req} in request body."
                    )
            # Check Request Body ---------------------------------------- Finish

            # Initialize Data Input ---------------------------------------- Start
            supplier_id = datas["supplier_id"]
            purchase_date = datas["purchase_date"]
            purchase_details = datas["purchase_details"]
            # Initialize Data Input ---------------------------------------- Finish

            # Data Validation ---------------------------------------- Start
            checker_result = purchase_validator(
                supplier_id,
                purchase_date,
                purchase_details,
                workshop_id
            )

            if len(checker_result) != 0:
                return defined_error(
                    checker_result,
                    "Defined Error",
                    499
                )
            # Data Validation ---------------------------------------- Finish

            # Check Workshop ---------------------------------------- Start
            workshop = Workshops.query.filter_by(
                id=workshop_id,
                is_delete=0
            ).first()

            if not workshop:
                return not_found(
                    "Workshop could not be found."
                )
            # Check Workshop ---------------------------------------- Finish

            # Check Purchase ---------------------------------------- Start
            purchase = Purchases.query.filter_by(
                id=id,
                workshop_id=workshop_id,
                is_delete=0
            ).first()

            if not purchase:
                return not_found(
                    "Purchase could not be found."
                )
            # Check Purchase ---------------------------------------- Finish

            # Check Supplier ---------------------------------------- Start
            supplier = Suppliers.query.filter_by(
                id=supplier_id,
                workshop_id=workshop_id,
                is_delete=0
            ).first()

            if not supplier:
                return not_found(
                    "Supplier could not be found."
                )
            # Check Supplier ---------------------------------------- Finish

            timestamp = current_timestamp()
            total = 0

            # Rollback Old Stock ---------------------------------------- Start
            for detail in purchase.purchase_details:
      
                  product = Products.query.filter_by(
                        id=detail.product_id,
                        workshop_id=workshop_id,
                        is_delete=0
                  ).first()

                  if product:

                        if product.stock < detail.quantity:
                              db.session.rollback()

                              return defined_error(
                              ["Product stock is insufficient."],
                              "Defined Error",
                              499
                              )

                        product.stock -= detail.quantity

                  db.session.delete(detail)

            db.session.flush()
            # Rollback Old Stock ---------------------------------------- Finish

            # Insert New Detail ---------------------------------------- Start
            for item in purchase_details:

                product = Products.query.filter_by(
                    id=item["product_id"],
                    workshop_id=workshop_id,
                    is_delete=0
                ).first()

                if not product:
                    db.session.rollback()

                    return not_found(
                        "Product could not be found."
                    )

                quantity = int(item["quantity"])
                unit_cost = int(item["unit_cost"])
                subtotal = quantity * unit_cost

                detail = PurchaseDetails(
                    purchase_id=purchase.id,
                    product_id=product.id,
                    quantity=quantity,
                    unit_cost=unit_cost,
                    subtotal=subtotal
                )

                db.session.add(detail)

                product.stock += quantity

                # Update harga beli terakhir
                product.purchase_price = unit_cost

                total += subtotal
            # Insert New Detail ---------------------------------------- Finish

            # Update Purchase ---------------------------------------- Start
            purchase.supplier_id = supplier_id
            purchase.purchase_date = purchase_date
            purchase.total = total
            purchase.updated_at = timestamp
            # Update Purchase ---------------------------------------- Finish

            # Commit ---------------------------------------- Start
            try:
                db.session.commit()

            except Exception as e:
                db.session.rollback()
                return parameter_error(str(e))
            # Commit ---------------------------------------- Finish

            # Return Response ========================================
            return success(
                status_code=200
            )

        except Exception as e:
            db.session.rollback()
            return bad_request(str(e))
    # UPDATE PURCHASE ============================================================ End

    # DELETE PURCHASE ============================================================ Begin
    def delete_purchase(user_role, workshop_id, id):
        try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                return authorization_error()
            
            subscription_access = subscription_validator(user_role, workshop_id)

            if not subscription_access:
                return subscription_required()
            # Access Validation ---------------------------------------- Finish

            # Check Workshop ---------------------------------------- Start
            workshop = Workshops.query.filter_by(
                id=workshop_id,
                is_delete=0
            ).first()

            if not workshop:
                return not_found(
                    "Workshop could not be found."
                )
            # Check Workshop ---------------------------------------- Finish

            # Check Purchase ---------------------------------------- Start
            purchase = Purchases.query.filter_by(
                id=id,
                workshop_id=workshop_id,
                is_delete=0
            ).first()

            if not purchase:
                return not_found(
                    "Purchase could not be found."
                )
            # Check Purchase ---------------------------------------- Finish

            timestamp = current_timestamp()

            # Rollback Stock ---------------------------------------- Start
            for detail in purchase.purchase_details:

                product = Products.query.filter_by(
                    id=detail.product_id,
                    workshop_id=workshop_id,
                    is_delete=0
                ).first()

                if product:
      
                  if product.stock < detail.quantity:
                        db.session.rollback()

                        return defined_error(
                              ["Product stock is insufficient."],
                              "Defined Error",
                              499
                        )

                  product.stock -= detail.quantity
            # Rollback Stock ---------------------------------------- Finish

            # Soft Delete Purchase ---------------------------------------- Start
            purchase.is_delete = 1
            purchase.deleted_at = timestamp
            purchase.updated_at = timestamp
            # Soft Delete Purchase ---------------------------------------- Finish

            # Commit ---------------------------------------- Start
            try:
                db.session.commit()

            except Exception as e:
                db.session.rollback()
                return parameter_error(str(e))
            # Commit ---------------------------------------- Finish

            # Return Response ========================================
            return success(
                status_code=200
            )

        except Exception as e:
            db.session.rollback()
            return bad_request(str(e))
    # DELETE PURCHASE ============================================================ End

    # IMPORT PURCHASE ============================================================ Begin
    def import_purchase(user_role, workshop_id, supplier_id, purchase_date, file):
      try:
            # Access Validation ---------------------------------------- Start
            access = role_validator(user_role)

            if not access:
                  return authorization_error()
            
            subscription_access = subscription_validator(user_role, workshop_id)

            if not subscription_access:
                return subscription_required()
            # Access Validation ---------------------------------------- Finish

            # Excel File Validation ---------------------------------------- Start
            checker_result = excel_file_validator(file)

            if len(checker_result) != 0:
                  return defined_error(
                  checker_result,
                  "Defined Error",
                  499
                  )
            # Excel File Validation ---------------------------------------- Finish

            # Read Workbook ---------------------------------------- Start
            workbook = load_workbook(file)
            worksheet = workbook.active
            # Read Workbook ---------------------------------------- Finish

            # Excel Content Validation ---------------------------------------- Start
            checker_result = purchase_excel_validator(
                  worksheet
            )

            if len(checker_result) != 0:
                  return defined_error(
                  checker_result,
                  "Defined Error",
                  499
                  )
            # Excel Content Validation ---------------------------------------- Finish

            # Initialize Purchase Detail ---------------------------------------- Start
            purchase_details = []
            # Initialize Purchase Detail ---------------------------------------- Finish

             # Read Excel ---------------------------------------- Start
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=2,
                    values_only=True
                ),
                start=2
            ):
                barcode, quantity, unit_cost = row

                # Ignore Empty Row
                if barcode in [None, ""] and quantity in [None, ""] and unit_cost in [None, ""]:
                    continue

                barcode = str(barcode).strip()

                product = Products.query.filter_by(
                    barcode=barcode,
                    workshop_id=workshop_id,
                    is_delete=0
                ).first()

                if not product:
                    return defined_error(
                        [
                            f"Barcode '{barcode}' pada baris {row_number} tidak ditemukan."
                        ],
                        "Defined Error",
                        499
                    )

                purchase_details.append({
                    "product_id": product.id,
                    "quantity": int(quantity),
                    "unit_cost": int(unit_cost)
                })
            # Read Excel ---------------------------------------- Finish

            # Initialize Data ---------------------------------------- Start
            datas = {
                  "supplier_id": supplier_id,
                  "purchase_date": purchase_date,
                  "purchase_details": purchase_details
            }
            # Initialize Data ---------------------------------------- Finish

            # Create Purchase ---------------------------------------- Start
            return PurchaseModels.create_purchase(
                  user_role,
                  workshop_id,
                  datas
            )
            # Create Purchase ---------------------------------------- Finish

      except Exception as e:
            return bad_request(str(e))
    # IMPORT PURCHASE ============================================================ End

# PURCHASE MODEL CLASS ============================================================ End